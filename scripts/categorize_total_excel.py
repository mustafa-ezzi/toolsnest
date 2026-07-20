"""Rebuild a clean categorized Excel from the Total .xlsm source.

Avoids corruption by writing a fresh .xlsx (data only, no VBA/images).
"""

from __future__ import annotations

import csv
import re
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

SRC = Path(r"D:\ToolsNest.tools\docs\updated total products.xlsm")
OUT = Path(r"D:\ToolsNest.tools\docs\updated-total-products-with-category.xlsx")
OUT_CSV = Path(r"D:\ToolsNest.tools\docs\updated-total-products-with-category.csv")


def infer_category(name: str, desc: str, sku: str) -> str:
    name_l = (name or "").lower().strip()
    desc_l = (desc or "").lower()
    sku_u = (sku or "").upper().strip()

    name_rules: list[tuple[str, list[str]]] = [
        (
            "Kitchen Appliances",
            [
                r"blender",
                r"food processor",
                r"mixer",
                r"kettle",
                r"toaster",
                r"juicer",
                r"coffee",
                r"chopper",
                r"cooker",
            ],
        ),
        (
            "Cleaning",
            [r"vacuum", r"pressure washer", r"steam clean", r"carpet clean", r"\bwasher\b"],
        ),
        (
            "Gardening & Outdoor",
            [
                r"lawn",
                r"hedge",
                r"grass",
                r"trimmer",
                r"leaf blower",
                r"chainsaw",
                r"garden",
                r"pruner",
            ],
        ),
        ("Welding & Soldering", [r"weld", r"solder"]),
        (
            "Automotive",
            [r"jump starter", r"tire inflator", r"car wash", r"automotive compressor"],
        ),
        (
            "Measuring & Layout",
            [
                r"^laser",
                r"laser level",
                r"laser distance",
                r"rangefinder",
                r"tape measure",
                r"spirit level",
                r"digital level",
                r"multimeter",
                r"detector",
                r"caliper",
            ],
        ),
        (
            "Safety",
            [
                r"glove",
                r"goggle",
                r"helmet",
                r"ear muff",
                r"safety glass",
                r"respirator",
                r"face shield",
            ],
        ),
        (
            "Lighting",
            [
                r"^led ",
                r"work light",
                r"flashlight",
                r"torch",
                r"flood light",
                r"camping light",
                r"headlamp",
                r"\blamp\b",
            ],
        ),
        (
            "Batteries & Chargers",
            [
                r"^battery",
                r"battery pack",
                r"^charger",
                r"battery and charger",
                r"lithium-ion battery",
            ],
        ),
        (
            "Storage & Accessories",
            [r"toolbox", r"tool box", r"tool bag", r"organizer", r"storage case"],
        ),
        (
            "Fasteners & Consumables",
            [
                r"cutting disc",
                r"flap disc",
                r"saw blade",
                r"drill bit set",
                r"bit set",
                r"sandpaper",
                r"abrasive",
            ],
        ),
        (
            "Hand Tools",
            [
                r"\bpliers?\b",
                r"\bspanner\b",
                r"adjustable wrench",
                r"combination wrench",
                r"socket set",
                r"ratchet",
                r"allen key",
                r"hex key",
                r"utility knife",
                r"crowbar",
                r"\baxe\b",
                r"mallet",
                r"hand saw",
                r"pipe wrench",
                r"vise grip",
                r"clamp",
                r"^screwdriver$",
                r"manual screwdriver",
            ],
        ),
        (
            "Power Tools",
            [
                r"impact drill",
                r"impact wrench",
                r"impact driver",
                r"angle grinder",
                r"rotary hammer",
                r"circular saw",
                r"jigsaw",
                r"reciprocating",
                r"demolition",
                r"hammer drill",
                r"polisher",
                r"sander",
                r"router",
                r"planer",
                r"miter saw",
                r"table saw",
                r"heat gun",
                r"die grinder",
                r"multi-?tool",
                r"oscillating",
                r"nailer",
                r"cordless",
                r"brushless",
                r"electric drill",
                r"screwdrive",
                r"compact brushless",
                r"grinder",
                r"drill",
            ],
        ),
        ("Electrical", [r"extension cord", r"wire stripper", r"voltage tester"]),
    ]

    for cat, patterns in name_rules:
        for p in patterns:
            if re.search(p, name_l):
                return cat

    if sku_u.startswith(("TBLI", "TCLI", "TFBLI", "TFCLI", "TFBCPM", "TBCL")):
        return "Batteries & Chargers"

    if sku_u.startswith(
        (
            "TDLI",
            "TIDLI",
            "TIWLI",
            "TAGLI",
            "TRHLI",
            "TAPLI",
            "TSLI",
            "TSDLI",
            "TMGLI",
            "TCSLI",
            "TJSLI",
            "TACLI",
            "THGLI",
            "TOSLI",
            "TCLLI",
            "TPSLI",
            "TBEK",
            "TIDV",
        )
    ):
        return "Power Tools"

    if re.search(r"brushless motor|no-load speed|impact rate|chuck capacity", desc_l):
        return "Power Tools"
    if re.search(r"lithium-ion .*battery|output current:.*a", desc_l) and "drill" not in name_l:
        return "Batteries & Chargers"

    return "Power Tools"


def main() -> None:
    # Read values only from source (ignore VBA / drawing corruption paths)
    src = openpyxl.load_workbook(SRC, data_only=True, keep_vba=False, read_only=True)
    ws_in = src["Offers"]

    rows: list[tuple] = []
    for i, row in enumerate(ws_in.iter_rows(min_row=3, values_only=True), start=3):
        sku = row[0] if row else None
        name = row[2] if row and len(row) > 2 else None
        desc = row[3] if row and len(row) > 3 else None
        price = row[4] if row and len(row) > 4 else None
        brand = row[5] if row and len(row) > 5 else None
        if not sku and not name:
            continue
        brand_out = str(brand).strip().upper() if brand else "TOTAL"
        cat = infer_category(str(name or ""), str(desc or ""), str(sku or ""))
        rows.append((sku, name, desc, price, brand_out, cat))
    src.close()

    # Fresh clean workbook (no images/VBA) — opens reliably in Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = [
        "SKU",
        "Product name",
        "Description & Features",
        "Selling Price",
        "Brand",
        "Category",
    ]
    header_fill = PatternFill("solid", fgColor="117076")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    counts: Counter[str] = Counter()
    for r, (sku, name, desc, price, brand, cat) in enumerate(rows, start=2):
        ws.cell(r, 1, sku)
        ws.cell(r, 2, name)
        ws.cell(r, 3, desc)
        ws.cell(r, 4, price)
        ws.cell(r, 5, brand)
        ws.cell(r, 6, cat)
        counts[cat] += 1

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(rows) + 1}"

    # Remove old corrupt file first
    if OUT.exists():
        OUT.unlink()
    wb.save(OUT)

    # Also write CSV fallback (always opens)
    with OUT_CSV.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)

    print(f"Saved Excel: {OUT}")
    print(f"Saved CSV:   {OUT_CSV}")
    print(f"Rows: {len(rows)}")
    for k, v in counts.most_common():
        print(f"  {v:4d}  {k}")


if __name__ == "__main__":
    main()
