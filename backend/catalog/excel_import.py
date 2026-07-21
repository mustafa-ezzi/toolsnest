"""Excel → products importer (SKU upsert, Cloudflare R2 images).

Supports:
- Clean data workbook (.xlsx/.csv-like sheets) with Brand + Category columns
- Optional images workbook (.xlsm/.xlsx) with embedded pictures matched by row/SKU
- Images embedded in the same data workbook (if present)
"""

from __future__ import annotations

import logging
import mimetypes
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import BinaryIO, Callable

from django.db import transaction
from django.utils.text import slugify
from PIL import Image as PilImage

from catalog.models import Brand, Category, Product, ProductImage
from media_upload.storage import make_object_key, r2_configured, upload_bytes

logger = logging.getLogger(__name__)

BRAND_COLORS = {
    "total": ("#117076", "#0B5559"),
    "ingco": ("#E30613", "#9B0000"),
    "makita": ("#003DA5", "#00A3E0"),
    "dewalt": ("#FFCC00", "#1D1D1D"),
    "stanley": ("#1D1D1D", "#FFD100"),
}

ProgressFn = Callable[[str], None]


@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    images_uploaded: int = 0
    images_skipped: int = 0
    images_failed: int = 0
    images_missing: int = 0
    storage: str = "none"
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def as_dict(self) -> dict:
        return {
            "created": self.created,
            "updated": self.updated,
            "skipped": self.skipped,
            "images_uploaded": self.images_uploaded,
            "images_skipped": self.images_skipped,
            "images_failed": self.images_failed,
            "images_missing": self.images_missing,
            "storage": self.storage,
            "errors": self.errors[:50],
            "warnings": self.warnings[:50],
            "error_count": len(self.errors),
            "warning_count": len(self.warnings),
        }


def cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_bool(value, default: bool = False) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in ("1", "true", "yes", "y", "on"):
        return True
    if s in ("0", "false", "no", "n", "off"):
        return False
    return default


def parse_price(value) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError, TypeError):
        return None


def get_or_create_brand(name: str) -> Brand:
    name = (name or "Total").strip()
    existing = Brand.objects.filter(name__iexact=name).first()
    if existing:
        return existing
    display = name.title() if name.isupper() else name
    primary, secondary = BRAND_COLORS.get(name.lower(), ("#117076", "#0B5559"))
    brand, _ = Brand.objects.get_or_create(
        name=display,
        defaults={
            "primary_color": primary,
            "secondary_color": secondary,
            "is_active": True,
        },
    )
    return brand


def get_or_create_category(name: str) -> Category | None:
    name = (name or "").strip()
    if not name:
        return None
    existing = Category.objects.filter(name__iexact=name).first()
    if existing:
        return existing
    category, _ = Category.objects.get_or_create(
        name=name,
        defaults={"is_active": True},
    )
    return category


def normalize_image_bytes(data: bytes) -> bytes:
    max_bytes = 5 * 1024 * 1024
    if data[:3] == b"\xff\xd8\xff" and len(data) <= max_bytes:
        return data
    if data[:8] == b"\x89PNG\r\n\x1a\n" and len(data) <= max_bytes:
        return data

    with PilImage.open(BytesIO(data)) as im:
        if im.mode in ("RGBA", "P", "LA"):
            background = PilImage.new("RGB", im.size, (255, 255, 255))
            rgba = im.convert("RGBA")
            background.paste(rgba, mask=rgba.split()[-1])
            im = background
        elif im.mode != "RGB":
            im = im.convert("RGB")

        quality = 85
        out = data
        while quality >= 50:
            buf = BytesIO()
            im.save(buf, format="JPEG", quality=quality, optimize=True)
            out = buf.getvalue()
            if len(out) <= max_bytes:
                return out
            quality -= 10
        return out


def upload_product_image(sku: str, data: bytes) -> str:
    try:
        data = normalize_image_bytes(data)
    except Exception:
        logger.exception("Image normalize failed for %s", sku)

    safe = slugify(sku) or "product"
    key = make_object_key("products", filename=f"{safe}.jpg", ext=".jpg")
    if data[:8] == b"\x89PNG\r\n\x1a\n":
        key = make_object_key("products", filename=f"{safe}.png", ext=".png")
    content_type = mimetypes.guess_type(key)[0] or "image/jpeg"
    url, _storage = upload_bytes(data, key=key, content_type=content_type)
    return url


def _header_map(headers: list[str]) -> dict[str, int]:
    return {h.lower().strip(): i for i, h in enumerate(headers)}


def _col(colmap: dict[str, int], *names: str) -> int | None:
    for name in names:
        if name.lower() in colmap:
            return colmap[name.lower()]
    return None


def _rows_from_header_and_iter(
    headers: list[str],
    rows_iter,
    *,
    limit: int | None,
    default_brand: str,
    default_category: str,
    default_stock: int,
    warnings: list[str],
) -> list[dict]:
    col = _header_map(headers)
    idx_sku = _col(col, "SKU", "TOTAL Item No.", "Item No.", "Item Number")
    idx_name = _col(col, "Product name", "Product Name", "Name")
    idx_price = _col(col, "Selling Price", "Price", "Sale Price")
    if idx_sku is None or idx_name is None or idx_price is None:
        raise ValueError(
            f"Missing required columns (need SKU, Product name, Selling Price). "
            f"Found: {headers}"
        )

    idx_desc = _col(col, "Description & Features", "Description", "Features")
    idx_brand = _col(col, "Brand")
    idx_category = _col(col, "Category")
    idx_stock = _col(col, "Stock", "Stock Qty", "Qty", "Quantity")
    idx_compare = _col(col, "Compare at price", "Compare At Price", "Was Price")
    idx_featured = _col(col, "Featured")
    idx_active = _col(col, "Active", "Is Active")

    out: list[dict] = []
    for raw in rows_iter:
        if limit is not None and len(out) >= limit:
            break
        # Support both sequences and dict-like already-mapped rows
        def cell(idx: int | None):
            if idx is None:
                return None
            if idx >= len(raw):
                return None
            return raw[idx]

        sku = cell_str(cell(idx_sku))
        if not sku:
            continue
        name = cell_str(cell(idx_name))
        if not name:
            warnings.append(f"Skip {sku}: empty name")
            continue
        price = parse_price(cell(idx_price))
        if price is None or price <= 0:
            warnings.append(f"Skip {sku}: bad price")
            continue

        desc = cell_str(cell(idx_desc)) if idx_desc is not None else ""
        brand = cell_str(cell(idx_brand)) if idx_brand is not None else ""
        brand = brand or default_brand
        category = cell_str(cell(idx_category)) if idx_category is not None else ""
        category = category or default_category

        stock = default_stock
        if idx_stock is not None:
            try:
                stock = int(Decimal(str(cell(idx_stock)).replace(",", "").strip()))
            except Exception:
                pass

        compare_at = parse_price(cell(idx_compare)) if idx_compare is not None else None
        featured = (
            parse_bool(cell(idx_featured), False) if idx_featured is not None else False
        )
        is_active = (
            parse_bool(cell(idx_active), True) if idx_active is not None else True
        )

        out.append(
            {
                "sku": sku,
                "name": name,
                "description": desc,
                "price": price,
                "brand": brand,
                "category": category,
                "stock_qty": max(0, stock),
                "compare_at_price": compare_at,
                "featured": featured,
                "is_active": is_active,
            }
        )
    return out


def load_product_rows_from_csv(
    path: Path | str | BinaryIO,
    *,
    limit: int | None = None,
    default_brand: str = "Total",
    default_category: str = "",
    default_stock: int = 10,
) -> tuple[list[dict], list[str]]:
    import csv
    from io import StringIO

    warnings: list[str] = []

    def parse(reader) -> list[dict]:
        try:
            headers = [str(h or "").strip() for h in next(reader)]
        except StopIteration as exc:
            raise ValueError("CSV is empty") from exc
        return _rows_from_header_and_iter(
            headers,
            reader,
            limit=limit,
            default_brand=default_brand,
            default_category=default_category,
            default_stock=default_stock,
            warnings=warnings,
        )

    if hasattr(path, "read"):
        raw = path.read()
        if isinstance(raw, bytes):
            text = raw.decode("utf-8-sig")
        else:
            text = str(raw)
        rows = parse(csv.reader(StringIO(text)))
        return rows, warnings

    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = parse(csv.reader(f))
    return rows, warnings

def load_product_rows_from_workbook(
    path: Path | str | BinaryIO,
    *,
    sheet_name: str | None = None,
    limit: int | None = None,
    default_brand: str = "Total",
    default_category: str = "",
    default_stock: int = 10,
) -> tuple[list[dict], list[str]]:
    """Return (rows, warnings). path may be Path or file-like."""
    # CSV shortcut
    name = ""
    if isinstance(path, (str, Path)):
        name = str(path).lower()
    elif hasattr(path, "name"):
        name = str(getattr(path, "name") or "").lower()
    if name.endswith(".csv"):
        return load_product_rows_from_csv(
            path,
            limit=limit,
            default_brand=default_brand,
            default_category=default_category,
            default_stock=default_stock,
        )

    import openpyxl

    warnings: list[str] = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            # Prefer Products, else first sheet
            if "Products" in wb.sheetnames:
                ws = wb["Products"]
            elif "Offers" in wb.sheetnames:
                ws = wb["Offers"]
            else:
                ws = wb[wb.sheetnames[0]]
            warnings.append(f"Using sheet '{ws.title}'")

        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(h or "").strip() for h in next(rows_iter)]
        except StopIteration as exc:
            raise ValueError("Workbook is empty") from exc

        # Skip blank header row variants (TOTAL Offers often has title row)
        if not any(headers) or (
            headers[0]
            and "item" not in headers[0].lower()
            and headers[0].lower() not in ("sku", "total item no.", "product name")
            and not _col(_header_map(headers), "SKU", "TOTAL Item No.", "Product name")
        ):
            # Try next row as headers (Offers sheet layout)
            try:
                headers2 = [str(h or "").strip() for h in next(rows_iter)]
                if _col(
                    _header_map(headers2),
                    "SKU",
                    "TOTAL Item No.",
                    "Product name",
                    "Product Name",
                ) is not None:
                    headers = headers2
                    warnings.append("Skipped title row; using row 2 as headers")
            except StopIteration:
                pass

        out = _rows_from_header_and_iter(
            headers,
            rows_iter,
            limit=limit,
            default_brand=default_brand,
            default_category=default_category,
            default_stock=default_stock,
            warnings=warnings,
        )
        return out, warnings
    finally:
        wb.close()


def extract_images_by_sku(
    path: Path | str | BinaryIO,
    *,
    sheet_name: str | None = None,
    sku_col: int = 1,
    data_start_row: int = 3,
) -> dict[str, bytes]:
    """Map SKU → raw image bytes from embedded drawings."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        elif "Offers" in wb.sheetnames:
            ws = wb["Offers"]
        elif "Products" in wb.sheetnames:
            ws = wb["Products"]
        else:
            ws = wb[wb.sheetnames[0]]

        sku_by_row: dict[int, str] = {}
        for r in range(data_start_row, (ws.max_row or data_start_row) + 1):
            sku = cell_str(ws.cell(r, sku_col).value)
            if sku:
                sku_by_row[r] = sku

        # Also map from row 2 if headers were on row 2 (data from 3)
        # Already covered by data_start_row=3 default for TOTAL sheet.

        by_sku: dict[str, bytes] = {}
        images = list(getattr(ws, "_images", []) or [])
        for img in images:
            anchor = getattr(img, "anchor", None)
            marker = getattr(anchor, "_from", None) if anchor is not None else None
            if marker is None:
                continue
            excel_row = int(marker.row) + 1
            sku = sku_by_row.get(excel_row)
            if not sku or sku in by_sku:
                continue
            try:
                data = img._data()
            except Exception:
                continue
            if data:
                by_sku[sku] = data
        return by_sku
    finally:
        wb.close()


def import_products(
    *,
    data_path: Path | str | BinaryIO,
    images_path: Path | str | BinaryIO | None = None,
    data_sheet: str | None = None,
    images_sheet: str | None = None,
    default_brand: str = "Total",
    default_category: str = "",
    default_stock: int = 10,
    limit: int | None = None,
    skip_images: bool = False,
    update_images: bool = False,
    dry_run: bool = False,
    progress: ProgressFn | None = None,
) -> ImportResult:
    result = ImportResult()
    log = progress or (lambda msg: None)

    rows, warn = load_product_rows_from_workbook(
        data_path,
        sheet_name=data_sheet,
        limit=limit,
        default_brand=default_brand,
        default_category=default_category,
        default_stock=default_stock,
    )
    result.warnings.extend(warn)
    log(f"Loaded {len(rows)} product rows")

    images_by_sku: dict[str, bytes] = {}
    if not skip_images:
        # Prefer dedicated images workbook; also try data workbook embeds
        sources: list[tuple[str, Path | str | BinaryIO | None, str | None]] = []
        if images_path is not None:
            sources.append(("images file", images_path, images_sheet))
        sources.append(("data file", data_path, images_sheet or data_sheet))

        for label, src, sheet in sources:
            if src is None:
                continue
            try:
                extracted = extract_images_by_sku(src, sheet_name=sheet)
                for sku, blob in extracted.items():
                    if sku not in images_by_sku:
                        images_by_sku[sku] = blob
                log(f"Extracted {len(extracted)} images from {label}")
            except Exception as exc:
                result.warnings.append(f"Could not extract images from {label}: {exc}")

        log(f"Matched images for {len(images_by_sku)} SKUs")

    if dry_run:
        result.storage = "dry-run"
        for row in rows:
            if Product.objects.filter(sku=row["sku"]).exists():
                result.updated += 1
            else:
                result.created += 1
            if not skip_images and row["sku"] not in images_by_sku:
                result.images_missing += 1
        return result

    result.storage = "r2" if r2_configured() else "local"
    log(f"Image storage: {result.storage}")

    for i, row in enumerate(rows, start=1):
        sku = row["sku"]
        try:
            brand = get_or_create_brand(row["brand"])
            category = get_or_create_category(row["category"])
        except Exception as exc:
            result.errors.append(f"{sku}: brand/category — {exc}")
            result.skipped += 1
            continue

        defaults = {
            "name": row["name"],
            "description": row["description"],
            "price": row["price"],
            "brand": brand,
            "category": category,
            "stock_qty": row["stock_qty"],
            "is_active": row["is_active"],
            "featured": row["featured"],
            "compare_at_price": row.get("compare_at_price"),
        }

        try:
            with transaction.atomic():
                product, was_created = Product.objects.update_or_create(
                    sku=sku, defaults=defaults
                )
        except Exception as exc:
            result.errors.append(f"{sku}: save failed — {exc}")
            result.skipped += 1
            continue

        if was_created:
            result.created += 1
        else:
            result.updated += 1

        if skip_images:
            if i % 50 == 0:
                log(f"Progress {i}/{len(rows)}…")
            continue

        img_bytes = images_by_sku.get(sku)
        if not img_bytes:
            result.images_missing += 1
            continue

        has_existing = product.images.exists()
        if has_existing and not update_images:
            result.images_skipped += 1
            continue

        try:
            url = upload_product_image(sku, img_bytes)
            if has_existing and update_images:
                product.images.all().delete()
            ProductImage.objects.create(
                product=product,
                url=url,
                alt=product.name,
                sort_order=0,
            )
            result.images_uploaded += 1
        except Exception as exc:
            result.images_failed += 1
            result.errors.append(f"{sku}: image upload — {exc}")
            logger.exception("Image upload failed for %s", sku)

        if i % 25 == 0 or i == len(rows):
            log(
                f"Progress {i}/{len(rows)} "
                f"(created={result.created} updated={result.updated} "
                f"imgs={result.images_uploaded})"
            )

    return result
